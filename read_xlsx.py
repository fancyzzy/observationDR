#!/usr/bin/env python3

'''
读取xlsx数据源
解析所有sheet名称
获取每个sheet的所有区间的名字和行范围
'''
import openpyxl
from datetime import datetime


class MyXlsx(object):
	def __init__(self, xlsx_path):
		self.path = xlsx_path
		print("load workbook: '%s'..."%(xlsx_path))
		self.wb = openpyxl.load_workbook(xlsx_path, data_only=True)
		print("load finsihed.")

		#获得所有的sheet页名单, 即观测项目
		self.sheets =  self.wb.sheetnames[:]

		#预先保存所有sheet的最大列数
		self.d_maxcol = {}
		self.d_maxrow = {}
		for sheet in self.sheets:
			self.d_maxcol[sheet] = len(tuple(self.wb[sheet].columns))
			self.d_maxrow[sheet] = len(tuple(self.wb[sheet].rows))

		#获取所有sheet的区间的行范围, 以字典形式为数据索引
		#注意表格格式，只适用于第一列是区间名，第二列是点号
		#{'sheet1':{'area1':(1,10), 'area2':(11,15),...}, 'sheet2':{'area4':(1,23)}}
		self.all_areas_row_range = self.get_all_sheets_areas_range()
		print("观测点范围:")
		for sheet in self.all_areas_row_range:
			print("{}: ".format(sheet))
			print(self.all_areas_row_range[sheet])
		#print(self.all_areas_row_range)

		#疑问，是否可以用地表沉降的区间作为全部区间?
		#不一定，通过对比set找到所有不同区间集合
		#获得'地表沉降'页的A列元素, 作为总区间汇总
		#['area1','area2','area3',...'area8']
		all_areas = []
		for sheet in self.all_areas_row_range.keys():
			all_areas.extend(self.all_areas_row_range[sheet].keys())
		all_areas = list(set(all_areas))

		self.areas = []
		try:
			sheet_name = '地表沉降'
			d_areas = self.all_areas_row_range[sheet_name]
			self.areas = list(d_areas.keys())
			if len(self.areas) < len(all_areas):
				print("DEBUG 有不在地表沉降范围内的区间:{}".\
					format(set(all_areas)-set(self.areas)))
				self.areas = all_areas
		except:
			print("没有地表沉降")
			self.areas = all_areas
		print("共有{}个区间:'{}'".format(len(self.areas),self.areas))
	##################__init__()##############################	


	def get_all_sheets_areas_range(self):
		'''
		获取所有sheet的area名和其观测点的行数范围
		用一个字典嵌套字典作为将来获取表信息的索引数据库，例如下:
		all_sheets_areas_range = {'sheet1':{'area1':(1,10), 'area2':(11,15),...}, \
		'sheet2':{'area4':(1,23)}}
		'''
		all_sheets_areas_range = {}
		sheet_areas_range = {}

		for sheet in self.sheets:
			#test one sheet
			#if sheet != '全站仪收敛(TBM)':
			#	continue
			sheet_areas_range = self.get_one_sheet_areas_range(sheet)
			if sheet_areas_range != None:
				all_sheets_areas_range[sheet] = sheet_areas_range
			else:
				all_sheets_areas_range[sheet] = {None:None}

		return all_sheets_areas_range
	#############get_all_sheets_areas_range()#################		


	def get_one_sheet_areas_range(self, sheet_name,target_col=1,from_row=1):
		'''
		获取一个sheet的所有区间(target_col=1)的行范围
		input:
		sheet_name 页名
		target_col 以哪一列为进准，根据初值列是否有值，判定找这个列的行范围
		说白了就是要获取合并单元格的行范围!

		output:
		sheet_areas_range = {'area1':(1,10), 'area2':(11,15),...'area4':(30,35)} 
		含义是{区间名:(起始行数,结束行数)

		'''
		#以初值列划定区间的行号范围
		init_col,_ = self.get_item_point(sheet_name, '初值', False)

		if not init_col:
			print("Error, {}没有'初值'列".format(sheet_name))
			return None

		sheet_areas_range = {}
		area_name = ''
		sheet = self.wb[sheet_name]
		start = 0
		start_count = False
		#多找10行,避免只有一个区间的表最后一行就是区间的最后，无法满足
		#三列都是空
		len_max_rows = len(tuple(sheet.rows)) + 1+ 10
		for i in range(from_row, len_max_rows):
			#表格格式注意, 区间必须是在A列, A列开始为空
			#以target_col+1是否有值来
			v_1_col = sheet.cell(row=i, column=target_col).value
			v_2_col = sheet.cell(row=i, column=target_col+1).value
			v_init = sheet.cell(row=i, column=init_col).value
			if v_1_col != None and (not start_count):
				area_name = v_1_col
				start = i
				start_count = True

			#发现新的area, 保存之前area的name,和上一行的行号i-1
			elif v_1_col != None and start_count:
				sheet_areas_range[area_name] = (start,i-1)
				#start 重新开始记录
				area_name = v_1_col
				start = i

			#最后一行结束以3列的值全为空，为结束，并且已经开始计数
			#表格格式注意, 观测点之间不能有空行
			elif v_1_col == None and v_2_col == None and start_count\
			and v_init == None:
				sheet_areas_range[area_name] = (start,i-1)
				#不支持空行
				break

			else:
				#继续寻找
				continue

		return sheet_areas_range
	#######get_one_sheet_areas_range()###########################


	def get_item_point(self, sheet, item, from_last_search = True):
		'''
		寻找某一个item的在某一个sheet的坐标
		返回列号和行号。
		注意，excel里行号列号从1开始
		input:
		sheet_name页名字段
		item查找的内容，可以使datetime类型
		from_last_search true:从右边最大列往第一列找，false: 反向
		output:
		列坐标,行坐标
		目前仅从1到5行搜索相关的域名
		'''
		start = 0
		end = 0
		step = 0
		if from_last_search:
			start = self.d_maxcol[sheet]
			end = 0
			step = -1
		else:
			start = 1
			end = self.d_maxcol[sheet]+1
			step = 1

		sh = self.wb[sheet]
		for col_number in range(start,end,step):
			#表格格式注意:日期类型code中是datetime.datetime, Excel中单元格选择date格式
			for row_number in range(1,5):
				if sh.cell(row_number,col_number).value and\
				 item == sh.cell(row_number,col_number).value:
					return col_number, row_number
		print("DEBUG 在'{}'中前五行都没有发现'{}'".format(sheet,item))
		return None, None
	#########get_item_point()##########################################


	def get_range_values(self, sheet, area_name, col):
		'''
		通过sheet，area和列坐标
		返回area这一列的所有值, 到一个列表[]
		列入返回1月1日这一列的衡山路站的测量值
		'''
		#print("sheet = {}, area_name = {}, col = {}".format(sheet,area_name,col))
		sh = self.wb[sheet]
		range_values = []

		#获取area的测量点行范围row_range
		start_row, end_row = self.all_areas_row_range[sheet][area_name]

		for i in range(start_row, end_row+1):
			#接受包含None值
			range_values.append(sh.cell(i, col).value)


		#建筑物倾斜的值每两个做差，然后把差值赋值回第一个，第二个设为None
		tmp_values = []
		ln = len(range_values)
		if '建筑物倾斜' in sheet:
			for i in range(ln-1):
				if i%2 != 0:
					continue
				curr_v = range_values[i]
				next_v = range_values[i+1]
				diff_v = None
				if next_v != None and curr_v != None:
					diff_v = next_v - curr_v
				tmp_values.append(diff_v)
				tmp_values.append(None)

			range_values = tmp_values
			

		return range_values
	#########get_values()######################################


	def get_avail_rows_values(self, sheet, rows, col, accept_none = False):
		'''
		input:
		rows 一个连续数字的列表比如[2,3,4,5]
		返回rows列表范围的
		有值的行的index列表和值列表
		output:
		返回有效值行数，返回这个rows区间的所有有效值
		'''
		def is_number(s):
			try:
				float(s)
				return True
			except ValueError:
				pass
			return False

		sh = self.wb[sheet]
		avail_rows = []
		avail_values = []

		for row_index in rows:
			s_value = sh.cell(row_index,col).value
			#print("DEBUG row:{},col:{},value='{}'".format(row_index,col,s_value))
			if s_value != None and is_number(s_value):
				avail_rows.append(row_index)
				avail_values.append(float(s_value))
			else:
				#其他非法输出，当做None值接受
				if accept_none:
					avail_rows.append(row_index)
					avail_values.append(None)

		return avail_rows,avail_values
	###########get_avail_rows_values()################################


	def get_value(self, sheet, row, col):
		'''
		获取该单元格值
		'''
		return self.wb[sheet].cell(row,col).value



if __name__ == '__main__':

	print("start")
	xlsx_path = r"C:\Users\tarzonz\Desktop\演示工程A\一二工区计算表2018.1.1_da.xlsx"
	#wrong path test
	#xlsx_path = r"C:\Users\tarzonz\Desktop\演示工程A\abc一二工区计算表2018.1.1.xlsx"

	my_xlsx = MyXlsx(xlsx_path)

	#测试获得所有列的数目
	#all_column = my_xlsx.wb['地表沉降'].columns
	#print("DEBUG all_row=",len(all_row))
	#print("DEBUG all_col=",len(tuple(all_column)))

	#print("DEBUG ground_sheet_areas=",my_xlsx.total_areas)



	#测试找某一个日期的列坐标
	ws = my_xlsx.wb['地表沉降']

	ss = '2018/1/1'
	sd = datetime.strptime(ss, '%Y/%m/%d')
	print("DEBUG sd=",sd)
	i,_ = my_xlsx.get_item_point('地表沉降', sd)
	print("i=",i)

	x = my_xlsx.get_value('地表沉降',2,8)
	print("x=",x)
	print("type(x)=",str(type(x)))
	print("datetime.datetime==type(x)",'datetime'in str(type(x)))

	'''
	#获取某一个单元格的值
	dd = ws['WE2'].value
	ddd = ws.cell(2, 603).value
	print("Debug ddd=",ddd)
	print("DEBUG type(ddd)=",type(ddd))
	'''
	print("main end")

